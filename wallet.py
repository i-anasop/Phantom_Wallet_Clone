import customtkinter as ctk
from PIL import Image
import os
import time
import threading
import datetime
import hashlib
import hmac
import secrets
import shutil
import qrcode
import tempfile
from cryptography.fernet import Fernet
from tkinter import filedialog
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==========================================
# Loading the Terms and Conditions
# ==========================================
def load_text_file(filename):
    """Load text content from data/ directory"""
    try:
        path = os.path.join("data", filename)
        if os.path.exists(path):
            with open(path, "r") as f:
                return f.read().strip()
    except:
        pass
    return ""

# ==========================================
# THEME
# ==========================================
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

COLOR_BG = "#0a0e27"  
COLOR_CARD = "#141a2e" 
COLOR_PRIMARY = "#9945ff" 
COLOR_PRIMARY_HOVER = "#7d35ff"  
COLOR_SECONDARY = "#1a1f3a" 
COLOR_TEXT = "#ffffff"
COLOR_TEXT_MUTED = "#8f9cb9" 
COLOR_SUCCESS = "#31a24c" 
COLOR_DANGER = "#ff6b6b"

FONT_HEADER = ("Inter", 24, "bold")
FONT_SUBHEADER = ("Inter", 18, "bold")
FONT_BODY = ("Inter", 14)
FONT_SMALL = ("Inter", 12)
FONT_MONO = ("JetBrains Mono", 14)

ASSETS_PATH = "assets"

if not os.path.exists(ASSETS_PATH):
    os.makedirs(ASSETS_PATH)

# ==========================================
# DSA Concepts
# ==========================================

class Node:
    def __init__(self, data):
        self.data = data
        self.next: Optional['Node'] = None

# DSA Concept: Linked List
class AuditLogLinkedList:
    def __init__(self):
        self.head: Optional[Node] = None
        self.tail: Optional[Node] = None

    def add_log(self, message):
        new_node = Node(message)
        if self.head is None:
            self.head = new_node
            self.tail = new_node
        else:
            if self.tail is not None:
                self.tail.next = new_node
            self.tail = new_node

        try:
            os.makedirs("admin", exist_ok=True)
            with open(os.path.join("admin", "audit_log.txt"), "a") as f:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                f.write(f"[{timestamp}] {message}\n")
        except Exception:
            pass

# DSA Concept: Stack (Undo Stack)
class UndoStack:
    def __init__(self):
        self.stack = []
    
    def push(self, transaction):
        self.stack.append(transaction)
    
    def pop(self):
        if self.stack:
            return self.stack.pop()
        return None
    
    def peek(self):
        if self.stack:
            return self.stack[-1]
        return None
    
    def size(self):
        return len(self.stack)

# DSA Concept: Queue (Transaction Queue)
class TransactionQueue:
    def __init__(self):
        self.queue = []
    
    def enqueue(self, transaction):
        self.queue.append(transaction)
    
    def dequeue(self):
        if self.queue:
            return self.queue.pop(0)
        return None
    
    def size(self):
        return len(self.queue)

# DSA Concept: Priority Queue (Priority Mempool)
class PriorityMempool:
    def __init__(self):
        self.mempool = []
    
    def add_transaction(self, priority, transaction):
        self.mempool.append({"priority": priority, "tx": transaction})
        self.mempool.sort(key=lambda x: x["priority"], reverse=True)
    
    def get_high_priority_tx(self):
        if self.mempool:
            return self.mempool.pop(0)
        return None
    
    def size(self):
        return len(self.mempool)

# ==========================================
# IS Concepts
# ==========================================

class ImageHelper:
    @staticmethod
    def crop_to_square(image_path, size=140):
        """Crop image to square and return PIL Image"""
        try:
            img = Image.open(image_path)
            # Get dimensions
            width, height = img.size
            # Determine crop box (center crop)
            if width == height:
                crop_size = width
            else:
                crop_size = min(width, height)
            left = (width - crop_size) // 2
            top = (height - crop_size) // 2
            right = left + crop_size
            bottom = top + crop_size
            # Crop and resize
            img = img.crop((left, top, right, bottom))
            img = img.resize((size, size), Image.Resampling.LANCZOS)
            return img
        except Exception as e:
            return None

# InfoSec Concept: AES Encryption, RSA Signatures, HMAC Authentication, SHA256 Hashing
class SecurityManager:
    def __init__(self):
        # Fernet (symmetric encryption)
        self.fernet_key = Fernet.generate_key()
        self.cipher = Fernet(self.fernet_key)
        
        # RSA keys for digital signatures
        from cryptography.hazmat.primitives.asymmetric import rsa
        self.private_key = rsa.generate_private_key(
            public_exponent=65537,
            key_size=2048
        )
        self.public_key = self.private_key.public_key()
        
        # HMAC secret key
        self.hmac_key = secrets.token_bytes(32)

    def sanitize_input(self, user_input):
        if not isinstance(user_input, str):
            return str(user_input)
        return user_input.replace("<", "").replace(">", "").strip()

    def validate_amount(self, amount):
        try:
            val = float(amount)
            return val > 0
        except ValueError:
            return False

    def generate_hash(self, data):
        """SHA256 Hash"""
        encoded = str(data).encode('utf-8')
        return hashlib.sha256(encoded).hexdigest()
    
    def generate_hmac(self, data):
        """HMAC authentication"""
        encoded = str(data).encode('utf-8')
        return hmac.new(self.hmac_key, encoded, hashlib.sha256).hexdigest()
    
    def encrypt_seed_phrase(self, seed_phrase):
        """AES Encryption (via Fernet) for seed phrase"""
        try:
            seed_str = " ".join(seed_phrase) if isinstance(seed_phrase, list) else seed_phrase
            encrypted = self.cipher.encrypt(seed_str.encode())
            return encrypted.decode()
        except Exception as e:
            return None
    
    def decrypt_seed_phrase(self, encrypted_seed):
        """Decrypt AES encrypted seed phrase"""
        try:
            decrypted = self.cipher.decrypt(encrypted_seed.encode())
            return decrypted.decode()
        except Exception as e:
            return None
    
    def sign_transaction(self, tx_data):
        """RSA Digital Signature for transaction"""
        from cryptography.hazmat.primitives import hashes
        from cryptography.hazmat.primitives.asymmetric import padding
        try:
            tx_str = str(tx_data).encode()
            signature = self.private_key.sign(
                tx_str,
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
            return signature.hex()[:64]  # Return first 64 chars
        except Exception as e:
            return None

    def generate_secure_id(self, length=16):
        """Secure random ID generation"""
        return secrets.token_hex(length // 2)

# ==========================================
# COIN METADATA
# ==========================================

COIN_METADATA = {
    "BTC": {"name": "Bitcoin", "icon": "bitcoin_btc_icon_transparent.png", "network": "BTC Taproot", "fee": 0.0001, "fee_coin": "BTC"},
    "ETH": {"name": "Ethereum", "icon": "ethereum_eth_icon_transparent.png", "network": "ETH Mainnet", "fee": 0.001, "fee_coin": "ETH"},
    "SOL": {"name": "Solana", "icon": "solana_sol_icon_transparent.png", "network": "Solana", "fee": 0.01, "fee_coin": "SOL"},
    "USDT": {"name": "USDT", "icon": "usdt_tether_icon_transparent.png", "network": "BSC Mainnet", "fee": 0.001, "fee_coin": "BNB", "fee_icon": "binance_bnb_icon_transparent.png"},
    "USDC": {"name": "USDC", "icon": "usdc_usd_coin_icon_transparent.png", "network": "BSC Mainnet", "fee": 0.001, "fee_coin": "BNB", "fee_icon": "binance_bnb_icon_transparent.png"},
    "POL": {"name": "Polygon", "icon": "polygon_matic_icon_transparent.png", "network": "Polygon Network", "fee": 0.2, "fee_coin": "MATIC"},
    "BNB": {"name": "BNB", "icon": "binance_bnb_icon_transparent.png", "network": "BSC Mainnet", "fee": 0.001, "fee_coin": "BNB"},
}

# ==========================================
# CORE WALLET LOGIC
# ==========================================

class WalletBackend:
    def __init__(self):
        self.history = []
        self.holdings = {}
        self.audit_log = AuditLogLinkedList()
        self.security = SecurityManager()
        self.undo_stack = UndoStack()
        self.transaction_queue = TransactionQueue()
        self.priority_mempool = PriorityMempool()

        self.wallet_data: Optional[dict] = None
        self.observers = []

        self.prices = {
            "BTC": 94230.50,
            "ETH": 3450.20,
            "SOL": 145.80,
            "USDT": 1.00,
            "USDC": 1.00,
            "MATIC": 0.85,
            "BNB": 580.40
        }

    def subscribe(self, callback):
        self.observers.append(callback)

    def notify(self):
        for callback in self.observers:
            callback()

    def create_wallet(self, seed_phrase=None):
        try:
            seed_words = ["apple", "river", "wind", "stone", "fire", "star", "moon", "sun", "tree", "ocean", "mountain", "cloud", "earth", "sky", "rain", "snow", "cloud", "beach", "ocean", "forest", "thunder", "silver", "golden", "bright", "swift", "calm"]

            # If importing existing wallet, use provided seed
            if seed_phrase:
                seed = seed_phrase.split()
            else:
                # Create new wallet with random seed
                seed = secrets.SystemRandom().sample(seed_words, 12)

            # DETERMINISTIC ADDRESS: Hash of seed phrase always produces same address
            seed_hash = hashlib.sha256(" ".join(seed).encode()).hexdigest()
            address = "0x" + seed_hash[:40]  # Use first 40 chars of hash

            self.wallet_data = {
                "name": "My Wallet 1",
                "address": address,
                "seed": seed,
                "balance": 0.0,
                "pfp": None
            }

            # DETERMINISTIC HOLDINGS: Use seed hash to generate reproducible amounts
            self.holdings = {}
            seed_int = int(seed_hash, 16)  # Convert hex to int for seeding

            for i, (token, price) in enumerate(self.prices.items()):
                # Use deterministic values based on seed - ENSURE NON-ZERO AMOUNTS
                if token in ["BTC", "ETH", "BNB"]:
                    # Generate amount between 0.5 and 5.0
                    amount = 0.5 + (((seed_int + i * 2000) % 450) / 100.0)
                else:
                    # Generate amount between 10 and 100
                    amount = float(10 + ((seed_int + i * 3000) % 90))

                self.holdings[token] = {"amount": amount, "price": price}

            self.recalculate_balance()
            self.log_wallet_creation(seed, address)
            self.audit_log.add_log(f"Wallet created: {address}")
            self.save_wallet_to_file(status="CREATED")
            self.notify()

        except Exception as e:
            print(f"Error creating wallet: {e}")

    def update_wallet_metadata(self, new_name, new_pfp_path=None):
        if not self.wallet_data:
            return

        clean_name = self.security.sanitize_input(new_name)
        if clean_name:
            self.wallet_data["name"] = clean_name
            # Update wallet name in wallets.txt
            self.update_wallet_in_file()

        if new_pfp_path:
            try:
                filename = os.path.basename(new_pfp_path)
                unique_name = f"{self.security.generate_secure_id(8)}_{filename}"
                dest_path = os.path.join(ASSETS_PATH, unique_name)
                shutil.copy(new_pfp_path, dest_path)
                self.wallet_data["pfp"] = dest_path
                # Save wallet with PFP to wallets.txt with CREATED status
                self.save_wallet_to_file_with_pfp(status="CREATED")
            except Exception as e:
                print(f"Failed to upload PFP: {e}")

        self.audit_log.add_log("Wallet metadata updated")
        self.notify()

    def update_wallet_in_file(self):
        """Update wallet name in wallets.txt when renamed"""
        try:
            os.makedirs("admin", exist_ok=True)
            filename = os.path.join("admin", "wallets.txt")

            if not os.path.exists(filename):
                return

            # Read all lines
            with open(filename, "r") as f:
                lines = f.readlines()

            # Update wallet name in matching address row
            with open(filename, "w") as f:
                for line in lines:
                    if self.wallet_data and self.wallet_data["address"] in line and "|" in line:
                        # Reconstruct the line with updated name
                        parts = line.split("|")
                        if len(parts) >= 6:
                            wallet_id = parts[0].strip()
                            timestamp = parts[2].strip()
                            address = parts[3].strip()
                            balance = f"${self.wallet_data['balance']:>13,.2f}"
                            seed_phrase = parts[5].strip()
                            line = f"{wallet_id:<5} | {self.wallet_data['name']:<15} | {timestamp:<20} | {address:<45} | {balance} | {seed_phrase}\n"
                    f.write(line)
        except Exception as e:
            pass

    def login(self, seed_phrase):
        clean_seed = self.security.sanitize_input(seed_phrase).strip()

        # Validate seed phrase format (must be 12 words)
        words = clean_seed.split()
        if len(words) != 12 or not all(isinstance(w, str) and len(w) > 0 for w in words):
            return False

        # Pass seed phrase to create_wallet for DETERMINISTIC generation
        self.create_wallet(seed_phrase=clean_seed)

        # Restore wallet name and PFP from wallets.txt if it exists (existing wallet)
        wallet_restored = False
        if self.wallet_data is not None:
            wallet_restored = self.restore_wallet_metadata()

        # Always save wallet info (with name, PFP, and status)
        if wallet_restored:
            # IMPORTED wallet - update with IMPORTED status
            self.save_wallet_to_file_with_pfp(status="IMPORTED")
        else:
            # NEW wallet - save with CREATED status
            self.save_wallet_to_file(status="CREATED")

        self.audit_log.add_log("User logged in with seed phrase")
        self.notify()
        return True

    def restore_wallet_metadata(self):
        """Restore wallet name and PFP from wallets.xlsx. Returns True if wallet found."""
        if not self.wallet_data:
            return False
        filename = os.path.join("admin", "wallets.xlsx")
        if not os.path.exists(filename):
            return False

        wb = None
        try:
            wb = load_workbook(filename)
            ws = wb.active

            # Find matching wallet by address (Address is column E = index 4)
            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[4].value == self.wallet_data.get("address"):
                    # Found matching wallet - RESTORE IT
                    if row[1].value:  # Wallet Name
                        self.wallet_data["name"] = row[1].value
                    # Check for PFP file in assets
                    if row[8].value:  # PFP Path
                        pfp_file = row[8].value
                        if os.path.exists(pfp_file):
                            self.wallet_data["pfp"] = pfp_file
                    return True  # Wallet was restored
            return False  # Wallet not found in file (new wallet)
        except Exception as e:
            return False
        finally:
            if wb:
                wb.close()

    def logout(self):
        self.wallet_data = None
        self.audit_log.add_log("User logged out")
        self.notify()
    
    def log_security_transaction(self, tx):
        """Log transaction with meaningful security data only"""
        os.makedirs("admin", exist_ok=True)
        with open(os.path.join("admin", "security_log.txt"), "a") as f:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            tx_type = tx.get("type", "unknown").upper()
            tx_id = tx.get("id", "unknown")
            tx_hash = tx.get("hash", "unknown")
            token = tx.get("token", "unknown")
            amount = tx.get("amount", 0)
            
            # Generate cryptographic signatures
            tx_data_str = f"{tx_id}{tx_type}{token}{amount}"
            hmac_sig = self.security.generate_hmac(tx_data_str)
            rsa_sig = self.security.sign_transaction(tx_data_str)
            
            f.write(f"\n[{timestamp}] {tx_type} | {token} {amount}")
            
            if tx_type == "SEND":
                recipient = tx.get('recipient', 'N/A')
                fee = tx.get('fee', 0)
                f.write(f" -> {recipient} (Fee: {fee})")
            
            f.write(f"\nTransaction ID: {tx_id}\n")
            f.write(f"SHA256 Hash: {tx_hash}\n")
            f.write(f"HMAC-SHA256: {hmac_sig}\n")
            f.write(f"RSA-2048 Sig: {rsa_sig}\n")
            f.write(f"{'-'*80}\n")
    
    def log_wallet_creation(self, seed_phrase, address):
        """Log wallet creation with security details"""
        os.makedirs("admin", exist_ok=True)
        with open(os.path.join("admin", "security_log.txt"), "a") as f:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Encrypt seed phrase for logging
            encrypted_seed = self.security.encrypt_seed_phrase(seed_phrase)
            seed_hash = self.security.generate_hash(" ".join(seed_phrase))
            
            f.write(f"\n[{timestamp}] WALLET CREATED\n")
            f.write(f"Address: {address}\n")
            f.write(f"Encrypted Seed (AES-256): {str(encrypted_seed)[:50]}...[PROTECTED]\n")
            f.write(f"Seed Hash (SHA256): {seed_hash}\n")
            f.write(f"{'-'*80}\n")

    def send_transaction_execute(self, recipient, amount, token):
        if not self.security.validate_amount(amount):
            return "Invalid amount"
        amount = float(amount)
        recipient = self.security.sanitize_input(recipient)

        # Get network fee
        fee = COIN_METADATA.get(token, {}).get("fee", 0)
        total_amount = amount + fee

        if token not in self.holdings or self.holdings[token]["amount"] < total_amount:
            return f"❌ Insufficient funds! You have {self.holdings[token]['amount']:.6g} {token}, but need {total_amount:.6g} {token} (including {fee} fee)"

        self.holdings[token]["amount"] -= total_amount
        self.recalculate_balance()

        tx = {
            "id": self.security.generate_secure_id(8),
            "type": "send",
            "token": token,
            "amount": amount,
            "fee": fee,
            "recipient": recipient,
            "time": datetime.datetime.now(),
            "status": "pending",
            "hash": self.security.generate_hash(f"{recipient}{amount}{token}")
        }

        self.history.insert(0, tx)
        
        # Hook up DSA structures
        self.undo_stack.push(tx)
        self.transaction_queue.enqueue(tx)
        self.priority_mempool.add_transaction(amount, tx)
        
        # Log security details
        self.log_security_transaction(tx)
        self.audit_log.add_log(f"Sent {amount} {token} to {recipient} with {fee} fee")
        self.notify()

        threading.Thread(target=self._confirm_tx, args=(tx,), daemon=True).start()

        return "✅ Transaction sent successfully!"

    def receive_transaction(self, amount, token):
        # BUG FIX: Ensure token exists before receiving
        if token not in self.holdings:
            self.holdings[token] = {"amount": 0.0, "price": self.prices.get(token, 1.0)}

        tx = {
            "id": self.security.generate_secure_id(8),
            "type": "receive",
            "token": token,
            "amount": amount,
            "time": datetime.datetime.now(),
            "status": "pending",
            "hash": self.security.generate_hash(f"receive{amount}{token}")
        }

        self.history.insert(0, tx)
        
        # Hook up DSA structures
        self.undo_stack.push(tx)
        self.transaction_queue.enqueue(tx)
        self.priority_mempool.add_transaction(amount, tx)
        
        # Log security details
        self.log_security_transaction(tx)
        self.audit_log.add_log(f"Received {amount} {token}")
        self.notify()
        threading.Thread(target=self._confirm_receive, args=(tx, amount, token), daemon=True).start()

    def _confirm_tx(self, tx):
        time.sleep(2)
        tx["status"] = "confirmed"
        self.notify()

    def _confirm_receive(self, tx, amount, token):
        time.sleep(2)
        tx["status"] = "confirmed"
        # BUG FIX: Verify token exists before updating
        if token in self.holdings:
            self.holdings[token]["amount"] += amount
            self.recalculate_balance()
        self.notify()

    def recalculate_balance(self):
        total = 0
        for token, data in self.holdings.items():
            total += data["amount"] * data["price"]
        if self.wallet_data:
            self.wallet_data["balance"] = total

    def save_wallet_to_file_with_pfp(self, status="CREATED"):
        """Save wallet with PFP path and status to Excel file using atomic write"""
        os.makedirs("admin", exist_ok=True)
        filename = os.path.join("admin", "wallets.xlsx")
        wb = None
        temp_fd = None
        try:
            # Load or create workbook
            if os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Wallets"
                # Add headers
                headers = ["ID", "Wallet Name", "Created", "Last Access", "Address", "Balance", "Status", "Seed Phrase", "PFP Path"]
                ws.append(headers)
                # Style header row
                header_fill = PatternFill(start_color="9945FF", end_color="9945FF", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Check if wallet already exists
            wallet_exists = False
            now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if self.wallet_data:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[4].value == self.wallet_data["address"]:
                        # Update existing wallet entry
                        row[1].value = self.wallet_data['name']
                        row[3].value = now
                        row[5].value = self.wallet_data.get('balance', 0)
                        row[6].value = status
                        row[8].value = self.wallet_data.get("pfp", "")
                        wallet_exists = True
                        break

            # If wallet doesn't exist, add new entry
            if not wallet_exists and self.wallet_data:
                wallet_id = len([row for row in ws.iter_rows(min_row=2)]) + 1
                created_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                seed_phrase = " ".join(self.wallet_data.get("seed", []))
                ws.append([
                    wallet_id,
                    self.wallet_data['name'],
                    created_date,
                    created_date,
                    self.wallet_data['address'],
                    self.wallet_data.get('balance', 0),
                    status,
                    seed_phrase,
                    self.wallet_data.get("pfp", "")
                ])

            # Use atomic write: save to temp file first, then rename
            temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir="admin")
            os.close(temp_fd)
            wb.save(temp_path)
            if os.path.exists(filename):
                os.remove(filename)
            os.rename(temp_path, filename)
        except Exception as e:
            print(f"Error saving wallet to Excel: {e}")
        finally:
            if wb:
                wb.close()

    def save_wallet_to_file(self, status="CREATED"):
        """Save wallet info to admin/wallets.xlsx in Excel format with status tracking using atomic write"""
        if not self.wallet_data:
            return
        os.makedirs("admin", exist_ok=True)
        filename = os.path.join("admin", "wallets.xlsx")
        wb = None
        temp_fd = None
        try:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Load or create workbook
            if os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Wallets"
                # Add headers
                headers = ["ID", "Wallet Name", "Created", "Last Access", "Address", "Balance", "Status", "Seed Phrase", "PFP Path"]
                ws.append(headers)
                # Style header row
                header_fill = PatternFill(start_color="9945FF", end_color="9945FF", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Get new wallet ID
            wallet_id = len([row for row in ws.iter_rows(min_row=2)]) + 1
            seed_phrase = " ".join(self.wallet_data.get("seed", []))
            address = self.wallet_data.get("address", "0x")
            balance = float(self.wallet_data.get("balance", 0.0))
            name = str(self.wallet_data.get("name", "My Wallet"))
            pfp_path = str(self.wallet_data.get("pfp", ""))

            # Append new wallet row
            ws.append([
                wallet_id,
                name,
                timestamp,
                timestamp,
                address,
                balance,
                status,
                seed_phrase,
                pfp_path
            ])

            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 50
            ws.column_dimensions['I'].width = 50

            # Use atomic write: save to temp file first, then rename
            temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir="admin")
            os.close(temp_fd)
            wb.save(temp_path)
            if os.path.exists(filename):
                os.remove(filename)
            os.rename(temp_path, filename)
        except Exception as e:
            print(f"Error saving wallet to Excel: {e}")
        finally:
            if wb:
                wb.close()

    # DSA Concept: Bubble Sort
    def sort_history_by_amount(self, descending=True):
        n = len(self.history)
        for i in range(n):
            for j in range(0, n - i - 1):
                if descending:
                    if self.history[j]['amount'] < self.history[j + 1]['amount']:
                        self.history[j], self.history[j + 1] = self.history[j + 1], self.history[j]
                else:
                    if self.history[j]['amount'] > self.history[j + 1]['amount']:
                        self.history[j], self.history[j + 1] = self.history[j + 1], self.history[j]
        self.notify()

backend = WalletBackend()

# ==========================================
# UX/UI
# ==========================================

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Phantom Simulator")
        self.geometry("480x850")
        self.resizable(False, True)
        self.configure(fg_color=COLOR_BG)
        try:
            self.iconbitmap("assets/icon.ico")
        except:
            pass  # Icon file not found, skip

        # Center window on screen
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (480 // 2)
        y = (self.winfo_screenheight() // 2) - (850 // 2)
        self.geometry(f"480x850+{x}+{y}")

        self.container = ctk.CTkFrame(self, fg_color="transparent")
        self.container.pack(fill="both", expand=True)

        self.current_view = None
        self.show_splash()

    def show_view(self, view_class, *args):
        if self.current_view:
            self.current_view.destroy()
        self.current_view = view_class(self.container, *args)
        self.current_view.pack(fill="both", expand=True)

    def show_splash(self):
        self.show_view(SplashView, self.show_landing)

    def show_landing(self):
        self.show_view(LandingView, self.show_terms, self.show_login)

    def show_terms(self):
        self.show_view(TermsView, self.show_seed, self.show_landing)

    def show_seed(self):
        backend.create_wallet()
        self.show_view(SeedView, self.show_dashboard)

    def show_login(self):
        self.show_view(LoginView, self.show_dashboard, self.show_landing)

    def show_dashboard(self):
        self.show_view(DashboardView, self.show_landing)


class SplashView(ctk.CTkFrame):
    def __init__(self, parent, next_callback):
        super().__init__(parent, fg_color=COLOR_BG)
        # Use pack with expand=True and fill both to properly center content
        center = ctk.CTkFrame(self, fg_color="transparent")
        center.pack(expand=True)

        img = Image.open("assets/phantom_logo.png")
        img = img.resize((150, 150), Image.Resampling.LANCZOS)
        ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(150, 150))
        ctk.CTkLabel(center, image=ctk_img, text="").pack(pady=20)
        self.logo_img = ctk_img  # keep reference to avoid garbage collection

        ctk.CTkLabel(center, text="PHANTOM", font=("Inter", 32, "bold"), text_color=COLOR_TEXT).pack()
        loading = ctk.CTkProgressBar(center, width=150, height=4, progress_color=COLOR_PRIMARY)
        loading.set(0)
        loading.pack(pady=20)
        loading.start()
        self.after(2500, next_callback)


class LandingView(ctk.CTkFrame):
    def __init__(self, parent, on_create, on_login):
        super().__init__(parent, fg_color=COLOR_BG)
        hero = ctk.CTkFrame(self, fg_color="transparent")
        hero.pack(expand=True)

        img = Image.open("assets/phantom_logo.png")
        img = img.resize((150, 150), Image.Resampling.LANCZOS)
        ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(150, 150))
        ctk.CTkLabel(hero, image=ctk_img, text="").pack(pady=20)
        self.logo_img = ctk_img  # keep reference to avoid garbage collection

        ctk.CTkLabel(hero, text="Phantom Simulator", font=FONT_HEADER, text_color=COLOR_TEXT).pack()
        ctk.CTkLabel(hero, text="The next generation of crypto\nwallet simulation.", font=FONT_BODY, text_color=COLOR_TEXT_MUTED).pack(pady=10)
        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(fill="x", padx=40, pady=40)
        ctk.CTkButton(btns, text="✨ Create New Wallet", height=55, corner_radius=14, fg_color=COLOR_PRIMARY, hover_color=COLOR_PRIMARY_HOVER, font=("Inter", 16, "bold"), command=on_create, text_color=COLOR_TEXT).pack(fill="x", pady=12)
        ctk.CTkButton(btns, text="📥 I already have a wallet", height=55, corner_radius=14, fg_color=COLOR_CARD, hover_color=COLOR_SECONDARY, font=("Inter", 15, "bold"), text_color=COLOR_PRIMARY, border_width=2, border_color=COLOR_PRIMARY, command=on_login).pack(fill="x", pady=5)


class TermsView(ctk.CTkFrame):
    def __init__(self, parent, on_agree, on_back):
        super().__init__(parent, fg_color=COLOR_BG)
        self.scrolled_to_end = False
        self.on_agree_callback = on_agree

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=20)
        ctk.CTkButton(header, text="← Back", width=80, fg_color=COLOR_CARD, hover_color=COLOR_SECONDARY, border_width=2, border_color=COLOR_PRIMARY, text_color=COLOR_PRIMARY, font=("Inter", 13, "bold"), command=on_back).pack(side="left")
        ctk.CTkLabel(header, text="Legal & Privacy", font=FONT_SUBHEADER).pack(side="left", padx=15)

        self.content = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.content.pack(fill="both", expand=True, padx=20)
        self.content.bind("<MouseWheel>", self.on_scroll)
        self.content.bind("<Button-4>", self.on_scroll)
        self.content.bind("<Button-5>", self.on_scroll)

        content = self.content

        warn = ctk.CTkFrame(content, fg_color="#450a0a", corner_radius=10)
        warn.pack(fill="x", pady=10)
        ctk.CTkLabel(warn, text="⚠️ Security Warning", text_color=COLOR_DANGER, font=("Inter", 14, "bold")).pack(anchor="w", padx=15, pady=(15,5))
        ctk.CTkLabel(warn, text="We cannot recover your funds if you lose your seed phrase.", text_color="#fecaca", wraplength=380, justify="left", font=FONT_SMALL).pack(anchor="w", padx=15, pady=(0,15))

        ctk.CTkLabel(content, text="Terms of Service", font=("Inter", 16, "bold")).pack(anchor="w", pady=(20,5))

        # Load terms from external file
        terms_text = load_text_file("TERMS_OF_SERVICE.txt")
        for section in terms_text.split("\n\n"):
            if section.strip():
                lines = section.split("\n")
                title = lines[0]
                body = "\n".join(lines[1:])
                ctk.CTkLabel(content, text=title, font=("Inter", 16, "bold")).pack(anchor="w", pady=(10, 5))
                ctk.CTkLabel(content, text=body, font=FONT_BODY, text_color=COLOR_TEXT_MUTED, justify="left", wraplength=400).pack(anchor="w", pady=(0, 10))

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.pack(fill="x", padx=20, pady=20)

        # Checkbox with Go to End button on right
        check_row = ctk.CTkFrame(footer, fg_color="transparent")
        check_row.pack(fill="x", pady=(0, 20))

        self.check_var = ctk.BooleanVar(value=False)
        self.checkbox = ctk.CTkCheckBox(check_row, text="✓ I understand and agree", variable=self.check_var, command=self.toggle_btn, fg_color=COLOR_PRIMARY, text_color=COLOR_TEXT, state="disabled")
        self.checkbox.pack(side="left", anchor="w")

        self.go_end_btn = ctk.CTkButton(check_row, text="↓", width=45, height=45, corner_radius=22, fg_color=COLOR_CARD, hover_color=COLOR_SECONDARY, border_width=2, border_color=COLOR_PRIMARY, text_color=COLOR_PRIMARY, font=("Arial", 20, "bold"), command=self.scroll_to_end)
        self.go_end_btn.pack(side="right", padx=(10, 0))

        self.continue_btn = ctk.CTkButton(footer, text="Continue →", height=55, corner_radius=14, fg_color=COLOR_SECONDARY, state="disabled", command=self.on_agree_callback, font=("Inter", 15, "bold"))
        self.continue_btn.pack(fill="x")

    def on_scroll(self, event):
        self.after(100, self.check_if_at_end)

    def check_if_at_end(self):
        try:
            # Get scroll position
            scroll_pos = self.content._parent_canvas.yview()
            if scroll_pos[1] >= 0.95:  # 95% scrolled down
                if not self.scrolled_to_end:
                    self.scrolled_to_end = True
                    self.checkbox.configure(state="normal")
                    self.go_end_btn.configure(fg_color=COLOR_PRIMARY, text_color=COLOR_TEXT, text="✓")
        except:
            pass

    def scroll_to_end(self):
        self.content._parent_canvas.yview_moveto(1.0)
        self.scrolled_to_end = True
        self.checkbox.configure(state="normal")
        self.go_end_btn.configure(fg_color=COLOR_PRIMARY, text_color=COLOR_TEXT, text="✓")

    def toggle_btn(self):
        state = "normal" if self.check_var.get() and self.scrolled_to_end else "disabled"
        color = COLOR_PRIMARY if self.check_var.get() and self.scrolled_to_end else COLOR_SECONDARY
        self.continue_btn.configure(state=state, fg_color=color)


class SeedView(ctk.CTkFrame):
    def __init__(self, parent, on_finish):
        super().__init__(parent, fg_color=COLOR_BG)
        self.on_finish = on_finish
        ctk.CTkLabel(self, text="Secret Recovery Phrase", font=FONT_HEADER).pack(pady=(40, 10))
        ctk.CTkLabel(self, text="Save these words in a safe place.", text_color=COLOR_TEXT_MUTED).pack(pady=(0, 30))
        self.seed_frame = ctk.CTkFrame(self, fg_color=COLOR_SECONDARY, corner_radius=15)
        self.seed_frame.pack(padx=20, fill="x")
        self.is_revealed = False
        self.render_words()
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=20)
        self.reveal_btn = ctk.CTkButton(btn_frame, text="👁 Reveal", fg_color=COLOR_CARD, hover_color=COLOR_SECONDARY, border_width=2, border_color=COLOR_PRIMARY, text_color=COLOR_PRIMARY, font=("Inter", 13, "bold"), command=self.toggle_reveal)
        self.reveal_btn.pack(side="left", padx=5)
        self.copy_btn = ctk.CTkButton(btn_frame, text="📋 Copy", fg_color=COLOR_PRIMARY, hover_color=COLOR_PRIMARY_HOVER, text_color=COLOR_TEXT, font=("Inter", 13, "bold"), command=self.copy_phrase)
        self.copy_btn.pack(side="left", padx=5)
        ctk.CTkButton(self, text="✓ I have saved my phrase", height=55, corner_radius=14, fg_color=COLOR_PRIMARY, hover_color=COLOR_PRIMARY_HOVER, font=("Inter", 15, "bold"), command=self.finish).pack(fill="x", side="bottom", padx=20, pady=40)

    def render_words(self):
        for w in self.seed_frame.winfo_children():
            w.destroy()
        if backend.wallet_data is None:
            return
        words = backend.wallet_data["seed"]
        for i, word in enumerate(words):
            row, col = i // 3, i % 3
            f = ctk.CTkFrame(self.seed_frame, fg_color="#000000", corner_radius=8)
            f.grid(row=row, column=col, padx=5, pady=5, sticky="ew")
            ctk.CTkLabel(f, text=f"{i+1}", text_color=COLOR_TEXT_MUTED, width=20).pack(side="left", padx=5)
            ctk.CTkLabel(f, text=word if self.is_revealed else "****", font=FONT_MONO).pack(side="left", padx=5)
        for i in range(3):
            self.seed_frame.grid_columnconfigure(i, weight=1)

    def toggle_reveal(self):
        self.is_revealed = not self.is_revealed
        self.reveal_btn.configure(text="👁 Hide Phrase" if self.is_revealed else "👁 Reveal Phrase")
        self.render_words()

    def copy_phrase(self):
        if backend.wallet_data:
            phrase = " ".join(backend.wallet_data["seed"])
            try:
                root = self.winfo_toplevel()
                root.clipboard_clear()
                root.clipboard_append(phrase)
                self.copy_btn.configure(text="✓ Copied!")
                self.after(2000, lambda: self.copy_btn.configure(text="📋 Copy"))
            except:
                pass

    def finish(self):
        self.on_finish()


class LoginView(ctk.CTkFrame):
    def __init__(self, parent, on_success, on_back):
        super().__init__(parent, fg_color=COLOR_BG)
        self.on_success = on_success
        self.word_entries = []

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=20)
        ctk.CTkButton(header, text="←", width=40, fg_color="transparent", command=on_back).pack(side="left")
        ctk.CTkLabel(header, text="Import Wallet", font=FONT_SUBHEADER).pack(side="left", padx=10)

        ctk.CTkLabel(self, text="Enter your 12-word seed phrase (paste or type):", font=("Inter", 12, "bold")).pack(anchor="w", padx=20, pady=(15,10))

        # 12 input boxes (4 rows x 3 cols)
        boxes_frame = ctk.CTkFrame(self, fg_color="transparent")
        boxes_frame.pack(padx=20, pady=10, fill="x")
        boxes_frame.grid_columnconfigure(0, weight=1)
        boxes_frame.grid_columnconfigure(1, weight=1)
        boxes_frame.grid_columnconfigure(2, weight=1)

        for i in range(12):
            row, col = i // 3, i % 3
            box_container = ctk.CTkFrame(boxes_frame, fg_color="transparent")
            box_container.grid(row=row, column=col, padx=5, pady=5, sticky="ew")

            entry = ctk.CTkEntry(box_container, placeholder_text=f"Word {i+1}", height=35, fg_color=COLOR_SECONDARY, border_width=1, border_color=COLOR_PRIMARY)
            entry.pack(fill="x")
            entry.bind("<KeyRelease>", lambda e, idx=i: self.on_word_change(idx, e))
            self.word_entries.append(entry)

        self.error_lbl = ctk.CTkLabel(self, text="", text_color=COLOR_DANGER, font=FONT_SMALL)
        self.error_lbl.pack(pady=5)

        ctk.CTkButton(self, text="Import Wallet", height=50, corner_radius=12, fg_color=COLOR_PRIMARY, command=self.do_login).pack(fill="x", side="bottom", padx=20, pady=40)

    def on_word_change(self, index, event):
        """Handle paste event - split pasted text into boxes"""
        entry = self.word_entries[index]
        text = entry.get().strip()

        if " " in text:  # Paste detected
            words = text.split()
            for i, word in enumerate(words):
                if index + i < 12:
                    self.word_entries[index + i].delete(0, "end")
                    self.word_entries[index + i].insert(0, word)

    def do_login(self):
        words = [entry.get().strip().lower() for entry in self.word_entries]
        # Remove empty words & validate exactly 12 non-empty words
        words = [w for w in words if w]
        if len(words) != 12:
            self.error_lbl.configure(text="❌ Please enter exactly 12 words")
            return
        if backend.login(" ".join(words)):
            self.on_success()
        else:
            self.error_lbl.configure(text="❌ Invalid seed phrase. Try again.")


class EditProfileWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Edit Profile")
        self.geometry("400x500")
        self.configure(fg_color=COLOR_BG)

        ctk.CTkLabel(self, text="Edit Profile", font=FONT_SUBHEADER).pack(pady=20)

        ctk.CTkLabel(self, text="Wallet Name", text_color=COLOR_TEXT_MUTED).pack(anchor="w", padx=20)
        self.name_entry = ctk.CTkEntry(self, height=40)
        if backend.wallet_data is not None:
            self.name_entry.insert(0, backend.wallet_data["name"])
        self.name_entry.pack(fill="x", padx=20, pady=(5, 20))

        ctk.CTkLabel(self, text="Profile Picture", text_color=COLOR_TEXT_MUTED).pack(anchor="w", padx=20)

        # Image preview container (centered)
        self.img_preview_frame = ctk.CTkFrame(self, fg_color=COLOR_CARD, corner_radius=10, width=150, height=150)
        self.img_preview_frame.pack(pady=15)
        self.img_preview_frame.pack_propagate(False)

        self.img_preview_label = ctk.CTkLabel(self.img_preview_frame, text="No Image\nSelected", text_color=COLOR_TEXT_MUTED, font=FONT_SMALL)
        self.img_preview_label.pack(expand=True)

        self.pfp_btn = ctk.CTkButton(self, text="Upload Image", fg_color=COLOR_SECONDARY, command=self.upload_pfp)
        self.pfp_btn.pack(fill="x", padx=20, pady=(5, 20))
        self.selected_pfp_path = None

        ctk.CTkButton(self, text="Save Changes", height=50, fg_color=COLOR_PRIMARY, command=self.save).pack(fill="x", padx=20, pady=20)

    def upload_pfp(self):
        filename = filedialog.askopenfilename(title="Select Profile Picture", filetypes=[("Images", "*.png *.jpg *.jpeg")])
        if filename:
            self.selected_pfp_path = filename
            try:
                img = ImageHelper.crop_to_square(filename, 140)
                if img:
                    self.ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(140, 140))
                    self.img_preview_label.configure(image=self.ctk_img, text="")
                else:
                    self.show_error_popup("Failed to process image")
            except Exception as e:
                self.show_error_popup(f"Failed to load image: {str(e)}")

    def show_error_popup(self, message):
        error_win = ctk.CTkToplevel(self)
        error_win.title("Error")
        error_win.geometry("300x150")
        error_win.configure(fg_color=COLOR_BG)
        ctk.CTkLabel(error_win, text="❌ Error", text_color=COLOR_DANGER, font=FONT_SUBHEADER).pack(pady=10)
        ctk.CTkLabel(error_win, text=message, text_color=COLOR_TEXT, font=FONT_BODY, wraplength=250).pack(pady=10, padx=10)
        ctk.CTkButton(error_win, text="OK", fg_color=COLOR_PRIMARY, command=error_win.destroy).pack(pady=10)

    def save(self):
        new_name = self.name_entry.get()
        backend.update_wallet_metadata(new_name, self.selected_pfp_path)
        self.destroy()


class DashboardView(ctk.CTkFrame):
    def __init__(self, parent, on_logout):
        super().__init__(parent, fg_color=COLOR_BG)
        self.on_logout = on_logout
        backend.subscribe(self.refresh)
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True)

        # TOP: Wallet Address (Copyable with Icon)
        addr_top = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD, corner_radius=10)
        addr_top.pack(fill="x", padx=20, pady=(20, 10))

        addr_top_inner = ctk.CTkFrame(addr_top, fg_color="transparent")
        addr_top_inner.pack(fill="x", padx=12, pady=8)

        ctk.CTkLabel(addr_top_inner, text="Wallet Address", font=FONT_SMALL, text_color=COLOR_TEXT_MUTED).pack(anchor="w")

        addr_row = ctk.CTkFrame(addr_top_inner, fg_color="transparent")
        addr_row.pack(fill="x", pady=(5, 0))

        self.addr_display = ctk.CTkLabel(addr_row, text="0x...", font=("Courier", 12, "bold"), text_color="white")
        self.addr_display.pack(side="left", fill="x", expand=True, padx=(0, 10))

        # Copy Icon Button (just emoji, no text)
        self.copy_icon_btn = ctk.CTkButton(addr_row, text="📋", width=30, height=30, font=("Arial", 14),
                                          fg_color=COLOR_SECONDARY, hover_color=COLOR_PRIMARY, 
                                          command=self.copy_address)
        self.copy_icon_btn.pack(side="right", padx=(10, 0))

        # Header
        header = ctk.CTkFrame(self.scroll, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=(10, 5))

        self.avatar_frame = ctk.CTkFrame(header, fg_color="transparent")
        self.avatar_frame.pack(side="left")

        ctk.CTkButton(header, text="Edit", width=40, height=24, font=FONT_SMALL, 
                      fg_color=COLOR_SECONDARY, command=lambda: EditProfileWindow(self)).pack(side="left", padx=5)

        self.name_lbl = ctk.CTkLabel(header, text="My Wallet", font=("Inter", 16, "bold"))
        self.name_lbl.pack(side="left", padx=5)

        ctk.CTkButton(header, text="Log Out", width=60, fg_color="transparent", text_color=COLOR_DANGER, command=self.do_logout).pack(side="right")

        # Balance - Improved UI
        bal_frame = ctk.CTkFrame(self.scroll, fg_color="transparent")
        bal_frame.pack(fill="x", pady=(30, 20), padx=20)
        self.balance_lbl = ctk.CTkLabel(bal_frame, text="$0.00", font=("Inter", 56, "bold"), text_color="white")
        self.balance_lbl.pack()
        ctk.CTkLabel(bal_frame, text="Pnl +$24.50 (1.2%)", text_color=COLOR_SUCCESS, font=("Inter", 12)).pack(pady=(5, 0))

        # Actions
        actions = ctk.CTkFrame(self.scroll, fg_color="transparent")
        actions.pack(fill="x", padx=20, pady=20)
        actions.grid_columnconfigure(0, weight=1)
        actions.grid_columnconfigure(1, weight=1)
        actions.grid_columnconfigure(2, weight=1)
        self.make_btn(actions, "Send", "send_arrow_icon.png", 0, lambda: SendWindow(self))
        self.make_btn(actions, "Receive", "receive_arrow_icon.png", 1, lambda: ReceiveWindow(self))
        self.make_btn(actions, "History", "history_clock_icon.png", 2, lambda: HistoryWindow(self))

        # Holdings Header
        ctk.CTkLabel(self.scroll, text="Holdings", font=FONT_SUBHEADER).pack(anchor="w", padx=20, pady=(20, 15))

        # Holdings Container
        self.holdings_frame = ctk.CTkFrame(self.scroll, fg_color="transparent")
        self.holdings_frame.pack(fill="x", padx=20, pady=(0, 20), ipady=5)

        self.refresh()

    def make_btn(self, parent, txt, icon_file, col, cmd):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.grid(row=0, column=col)

        btn_image = None
        emoji_fallback = "❓"

        # Map file names to emoji fallbacks
        emoji_map = {
            "send_arrow_icon.png": "↗",
            "receive_arrow_icon.png": "↙",
            "history_clock_icon.png": "🕐",
            "📚": "📚"
        }
        emoji_fallback = emoji_map.get(icon_file, "❓")

        # Try to load image file
        if icon_file not in emoji_map or icon_file != "📚":
            icon_path = os.path.join(ASSETS_PATH, icon_file)
            if os.path.exists(icon_path):
                try:
                    btn_image = ctk.CTkImage(Image.open(icon_path), size=(20, 20))
                except:
                    pass

        btn_kwargs = {"width": 60, "height": 60, "corner_radius": 30, 
                      "fg_color": COLOR_SECONDARY, "hover_color": COLOR_PRIMARY, "command": cmd}

        if btn_image:
            ctk.CTkButton(f, text="", image=btn_image, **btn_kwargs).pack()
        else:
            # Use emoji fallback if image not found
            ctk.CTkButton(f, text=emoji_fallback, font=("Arial", 24), **btn_kwargs).pack()

        ctk.CTkLabel(f, text=txt, font=FONT_SMALL, text_color=COLOR_TEXT_MUTED).pack(pady=5)

    def copy_address(self):
        if backend.wallet_data:
            self.master.clipboard_clear()
            self.master.clipboard_append(backend.wallet_data["address"])
            self.copy_icon_btn.configure(text="✓")
            self.after(2000, lambda: self.copy_icon_btn.configure(text="📋"))

    def refresh(self):
        if backend.wallet_data is None:
            return

        try:
            # Check if widgets still exist before updating
            if not self.winfo_exists():
                return

            self.name_lbl.configure(text=backend.wallet_data["name"])
            self.addr_display.configure(text=backend.wallet_data["address"])
            self.balance_lbl.configure(text=f"${backend.wallet_data['balance']:,.2f}")
        except:
            return

        try:
            # Update Avatar
            for w in self.avatar_frame.winfo_children():
                w.destroy()
            if backend.wallet_data.get("pfp"):
                try:
                    img = ImageHelper.crop_to_square(backend.wallet_data["pfp"], 40)
                    if img:
                        ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=(40, 40))
                        img_label = ctk.CTkLabel(self.avatar_frame, text="", image=ctk_img, corner_radius=20, width=40, height=40, fg_color="transparent")
                        img_label.pack()
                    else:
                        ctk.CTkLabel(self.avatar_frame, text="W", width=40, height=40, fg_color=COLOR_PRIMARY, corner_radius=20, text_color="white").pack()
                except:
                    ctk.CTkLabel(self.avatar_frame, text="W", width=40, height=40, fg_color=COLOR_PRIMARY, corner_radius=20, text_color="white").pack()
            else:
                ctk.CTkLabel(self.avatar_frame, text="W", width=40, height=40, fg_color=COLOR_PRIMARY, 
                            corner_radius=20, text_color="white", font=("Arial", 18, "bold")).pack()

            # Clear and refresh holdings
            for w in self.holdings_frame.winfo_children():
                w.destroy()
        except:
            return

        try:
            sorted_holdings = sorted(backend.holdings.items(), key=lambda x: x[1]["amount"] * x[1]["price"], reverse=True)

            for token, data in sorted_holdings:
                if data["amount"] <= 0:
                    continue

                # Main holding row
                row = ctk.CTkFrame(self.holdings_frame, fg_color=COLOR_CARD, corner_radius=12)
                row.pack(fill="x", pady=10, ipady=15, padx=0)

                # Left: Icon + Name
                left_frame = ctk.CTkFrame(row, fg_color="transparent")
                left_frame.pack(side="left", padx=15, fill="y")

                icon_file = COIN_METADATA.get(token, {}).get("icon", "bitcoin_btc_icon_transparent.png")
                icon_path = os.path.join(ASSETS_PATH, icon_file)

                try:
                    if os.path.exists(icon_path):
                        img = ctk.CTkImage(Image.open(icon_path), size=(35, 35))
                        ctk.CTkLabel(left_frame, text="", image=img).pack()
                except:
                    pass

                info_frame = ctk.CTkFrame(row, fg_color="transparent")
                info_frame.pack(side="left", padx=10, fill="y")

                coin_name = COIN_METADATA.get(token, {}).get("name", token)
                ctk.CTkLabel(info_frame, text=coin_name, font=("Inter", 14, "bold")).pack(anchor="w")
                ctk.CTkLabel(info_frame, text=f"${data['price']:,.2f}", text_color=COLOR_TEXT_MUTED, font=FONT_SMALL).pack(anchor="w")

                # Right: Values
                right_frame = ctk.CTkFrame(row, fg_color="transparent")
                right_frame.pack(side="right", padx=15, fill="y")

                ctk.CTkLabel(right_frame, text=f"${(data['amount']*data['price']):,.2f}", font=("Inter", 14, "bold")).pack(anchor="e")
                ctk.CTkLabel(right_frame, text=f"{data['amount']:,.6g} {token}", text_color=COLOR_TEXT_MUTED, font=FONT_SMALL).pack(anchor="e")
        except:
            pass

    def do_logout(self):
        backend.logout()
        self.on_logout()

class SendWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.geometry("450x600")
        self.configure(fg_color=COLOR_BG)
        ctk.CTkLabel(self, text="Send Assets", font=FONT_SUBHEADER, text_color=COLOR_TEXT).pack(pady=20)

        self.addr = ctk.CTkEntry(self, placeholder_text="Recipient Address", height=40, 
                                 fg_color=COLOR_CARD, border_color=COLOR_PRIMARY, border_width=1,
                                 text_color=COLOR_TEXT, placeholder_text_color=COLOR_TEXT_MUTED)
        self.addr.pack(fill="x", padx=20, pady=10)

        self.amt = ctk.CTkEntry(self, placeholder_text="Amount", height=40,
                               fg_color=COLOR_CARD, border_color=COLOR_PRIMARY, border_width=1,
                               text_color=COLOR_TEXT, placeholder_text_color=COLOR_TEXT_MUTED)
        self.amt.pack(fill="x", padx=20, pady=10)

        available_tokens = [t for t, d in backend.holdings.items() if d["amount"] > 0]
        if not available_tokens:
            available_tokens = ["SOL"]

        self.token = ctk.CTkComboBox(self, values=available_tokens, height=40,
                                    fg_color=COLOR_CARD, border_color=COLOR_PRIMARY, border_width=1,
                                    text_color=COLOR_TEXT, button_color=COLOR_PRIMARY, dropdown_fg_color=COLOR_CARD)
        self.token.set(available_tokens[0])
        self.token.pack(fill="x", padx=20, pady=10)

        # Network Fee Display - INITIALLY HIDDEN
        self.fee_frame = ctk.CTkFrame(self, fg_color=COLOR_CARD, corner_radius=10)
        self.fee_frame.pack(fill="x", padx=20, pady=15)
        self.fee_frame.pack_forget()  # Hide initially

        self.fee_label = ctk.CTkLabel(self.fee_frame, text="", text_color=COLOR_TEXT_MUTED, font=FONT_SMALL, wraplength=350)
        self.fee_label.pack(padx=15, pady=10, anchor="w")

        # Status/Error message display
        self.status_label = ctk.CTkLabel(self, text="", text_color=COLOR_TEXT_MUTED, font=FONT_SMALL, wraplength=350)
        self.status_label.pack(pady=15, padx=20)

        # Confirm button (first step)
        self.confirm_btn = ctk.CTkButton(self, text="Review Transaction", height=50, fg_color=COLOR_PRIMARY, 
                                        hover_color=COLOR_PRIMARY_HOVER, text_color=COLOR_TEXT, command=self.confirm_click)
        self.confirm_btn.pack(fill="x", padx=20, pady=10)

        # Send button (second step, initially hidden)
        self.send_btn = ctk.CTkButton(self, text="Send Now", height=50, fg_color=COLOR_PRIMARY, 
                                     hover_color=COLOR_PRIMARY_HOVER, text_color=COLOR_TEXT, command=self.send_click)
        self.send_btn.pack_forget()

        # Cancel button (second step, initially hidden)
        self.cancel_btn = ctk.CTkButton(self, text="Cancel", height=50, fg_color=COLOR_SECONDARY, 
                                       hover_color=COLOR_CARD, text_color=COLOR_TEXT, command=self.reset_form)
        self.cancel_btn.pack_forget()

    def confirm_click(self):
        """Show fee details when user clicks Confirm"""
        addr = self.addr.get()
        amt = self.amt.get()
        token = self.token.get()

        if not addr:
            self.show_message("❌ Please enter recipient address", COLOR_DANGER)
            return
        if not amt:
            self.show_message("❌ Please enter amount", COLOR_DANGER)
            return

        # Validate amount format
        try:
            amount_float = float(amt)
            if amount_float <= 0:
                self.show_message("❌ Amount must be greater than 0", COLOR_DANGER)
                return
        except ValueError:
            self.show_message("❌ Invalid amount format", COLOR_DANGER)
            return

        # CHECK INSUFFICIENT FUNDS BEFORE SHOWING FEE
        metadata = COIN_METADATA.get(token, {})
        fee = metadata.get("fee", 0)
        total_needed = amount_float + fee
        current_balance = backend.holdings.get(token, {}).get("amount", 0)

        if current_balance < total_needed:
            self.show_message(f"❌ Insufficient funds! You have {current_balance:.6g} {token}, need {total_needed:.6g} {token}", COLOR_DANGER)
            return

        # Clear fee frame and rebuild with coin logo
        for w in self.fee_frame.winfo_children():
            w.destroy()

        # Show fee details with coin logo
        metadata = COIN_METADATA.get(token, {})
        network = metadata.get("network", "Unknown Network")
        fee = metadata.get("fee", 0)
        fee_coin = metadata.get("fee_coin", token)
        # Use fee_icon if specified (e.g., BNB icon for USDT/USDC fees), else use token icon
        fee_icon_file = metadata.get("fee_icon", metadata.get("icon", ""))

        # Create header with network name
        header = ctk.CTkLabel(self.fee_frame, text=f"🔗 {network}", font=("Inter", 12, "bold"), text_color=COLOR_TEXT)
        header.pack(anchor="w", padx=15, pady=(10, 5))

        # Create fee row with coin logo
        fee_row = ctk.CTkFrame(self.fee_frame, fg_color="transparent")
        fee_row.pack(anchor="w", padx=15, pady=(5, 10), fill="x")

        # Try to load fee coin logo
        try:
            icon_path = os.path.join(ASSETS_PATH, fee_icon_file)
            if os.path.exists(icon_path):
                img = ctk.CTkImage(Image.open(icon_path), size=(24, 24))
                logo_label = ctk.CTkLabel(fee_row, text="", image=img)
                logo_label.pack(side="left", padx=(0, 10))
        except:
            pass

        # Fee amount label
        fee_label = ctk.CTkLabel(fee_row, text=f"Network Fee: {fee} {fee_coin}", font=("Inter", 11), text_color=COLOR_TEXT_MUTED)
        fee_label.pack(side="left")

        self.fee_frame.pack(fill="x", padx=20, pady=15)

        # Switch buttons
        self.confirm_btn.pack_forget()
        self.send_btn.pack(fill="x", padx=20, pady=10)
        self.cancel_btn.pack(fill="x", padx=20, pady=10)

        self.show_message("✓ Ready to send. Click 'Send Now' to confirm.", COLOR_SUCCESS)

    def reset_form(self):
        """Hide fee details and go back to initial state"""
        self.fee_frame.pack_forget()
        self.send_btn.pack_forget()
        self.cancel_btn.pack_forget()
        self.confirm_btn.pack(fill="x", padx=20, pady=10)
        self.status_label.configure(text="")

    def send_click(self):
        addr = self.addr.get()
        amt = self.amt.get()
        token = self.token.get()

        result = backend.send_transaction_execute(addr, amt, token)
        if result:
            if result.startswith("✅"):
                self.show_message(result, COLOR_SUCCESS)
                self.after(1500, self.destroy)
            else:
                self.show_message(result, COLOR_DANGER)
                self.reset_form()

    def show_message(self, message, color):
        self.status_label.configure(text=message, text_color=color)
        # Animate the message - fade in effect by opacity
        self.animate_message(0)

    def animate_message(self, step):
        """Smooth animation for message appearance"""
        if step < 100:
            # Gradually show the message (you can enhance this with actual opacity if needed)
            self.after(10, lambda: self.animate_message(step + 10))

class ReceiveWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Receive")
        self.geometry("480x720")
        self.configure(fg_color=COLOR_BG)

        ctk.CTkLabel(self, text="Receive Assets", font=FONT_SUBHEADER, text_color=COLOR_TEXT).pack(pady=20)

        # QR Code Container
        if backend.wallet_data is not None:
            try:
                qr_data = f"phantom:{backend.wallet_data['address']}"
                qr_img = qrcode.make(qr_data)
                qr_img = qr_img.resize((250, 250), Image.Resampling.LANCZOS)
                qr_ctk = ctk.CTkImage(light_image=qr_img, dark_image=qr_img, size=(250, 250))

                qr_container = ctk.CTkFrame(self, fg_color=COLOR_CARD, corner_radius=15, border_width=2, border_color=COLOR_PRIMARY)
                qr_container.pack(pady=20, padx=20)

                qr_label = ctk.CTkLabel(qr_container, text="", image=qr_ctk, fg_color=COLOR_CARD)
                qr_label.pack(padx=10, pady=10)
            except Exception as e:
                ctk.CTkLabel(self, text=f"QR Error: {str(e)}", text_color=COLOR_DANGER).pack(pady=20)

        # Address display - SINGLE LINE WITH COPY ICON
        if backend.wallet_data is not None:
            addr_container = ctk.CTkFrame(self, fg_color="transparent")
            addr_container.pack(pady=15, padx=20, fill="x")

            addr_row = ctk.CTkFrame(addr_container, fg_color="transparent")
            addr_row.pack(fill="x")

            self.addr_display = ctk.CTkLabel(addr_row, text=backend.wallet_data["address"], font=("Courier", 11, "bold"), text_color="white")
            self.addr_display.pack(side="left", fill="x", expand=True, padx=(0, 10))

            # Copy Icon Button
            self.copy_btn = ctk.CTkButton(addr_row, text="📋", width=30, height=30, font=("Arial", 14),
                                         fg_color=COLOR_SECONDARY, hover_color=COLOR_PRIMARY, 
                                         command=self.copy_address)
            self.copy_btn.pack(side="right", padx=(10, 0))

        # Token and Amount Selection
        selection_frame = ctk.CTkFrame(self, fg_color="transparent")
        selection_frame.pack(fill="x", padx=20, pady=15)

        ctk.CTkLabel(selection_frame, text="Receive Token", font=("Inter", 11, "bold"), text_color=COLOR_TEXT_MUTED).pack(anchor="w")
        all_tokens = list(backend.holdings.keys())
        self.token_combo = ctk.CTkComboBox(selection_frame, values=all_tokens, height=40,
                                          fg_color=COLOR_CARD, border_color=COLOR_PRIMARY, border_width=1,
                                          text_color=COLOR_TEXT, button_color=COLOR_PRIMARY, dropdown_fg_color=COLOR_CARD)
        self.token_combo.set(all_tokens[0] if all_tokens else "SOL")
        self.token_combo.pack(fill="x", pady=(5, 15))

        ctk.CTkLabel(selection_frame, text="Amount", font=("Inter", 11, "bold"), text_color=COLOR_TEXT_MUTED).pack(anchor="w")
        self.amount_entry = ctk.CTkEntry(selection_frame, placeholder_text="Enter amount (e.g., 1.5)", height=40,
                                        fg_color=COLOR_CARD, border_color=COLOR_PRIMARY, border_width=1,
                                        text_color=COLOR_TEXT, placeholder_text_color=COLOR_TEXT_MUTED)
        self.amount_entry.pack(fill="x", pady=(5, 5))

        ctk.CTkButton(self, text="Simulate Receive", height=50, fg_color=COLOR_PRIMARY, 
                     hover_color=COLOR_PRIMARY_HOVER, text_color=COLOR_TEXT, command=self.receive_click).pack(fill="x", padx=20, pady=20)

    def copy_address(self):
        """Copy wallet address to clipboard"""
        if backend.wallet_data:
            self.master.clipboard_clear()
            self.master.clipboard_append(backend.wallet_data["address"])
            self.copy_btn.configure(text="✓")
            self.after(2000, lambda: self.copy_btn.configure(text="📋"))

    def receive_click(self):
        token = self.token_combo.get()
        amount_str = self.amount_entry.get()

        if not amount_str:
            return

        try:
            amount = float(amount_str)
            if amount <= 0:
                return
            backend.receive_transaction(amount, token)
            self.destroy()
        except ValueError:
            return

class HistoryWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Activity History")
        self.geometry("480x750")
        self.configure(fg_color=COLOR_BG)
        self.sort_descending = True  # Track sort direction

        # Header section with title and controls
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=20)
        
        ctk.CTkLabel(header, text="Activity", font=FONT_HEADER, text_color=COLOR_TEXT).pack(anchor="w", pady=(0, 15))

        # Sort buttons row
        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 10))
        
        ctk.CTkButton(btn_frame, text="📊 Sort by Amount", height=35, fg_color=COLOR_PRIMARY, 
                     hover_color=COLOR_PRIMARY_HOVER, text_color=COLOR_TEXT, command=self.toggle_sort).pack(side="left", padx=(0, 5), fill="x", expand=True)
        ctk.CTkButton(btn_frame, text="🆕 By Time", height=35, fg_color=COLOR_SECONDARY, 
                     text_color=COLOR_TEXT, command=self.sort_by_time).pack(side="left", padx=(5, 0), fill="x", expand=True)

        # Transaction list
        self.scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self.scroll.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        self.refresh()

    def toggle_sort(self):
        self.sort_descending = not self.sort_descending
        backend.sort_history_by_amount(self.sort_descending)
        self.refresh()
    
    def sort_by_time(self):
        # Already sorted by time on creation, just refresh
        self.refresh()

    def refresh(self):
        for w in self.scroll.winfo_children():
            w.destroy()

        if not backend.history:
            ctk.CTkLabel(self.scroll, text="No transactions yet", text_color=COLOR_TEXT_MUTED, font=FONT_BODY).pack(pady=30)
            return

        for idx, tx in enumerate(backend.history):
            # Main transaction card
            row = ctk.CTkFrame(self.scroll, fg_color=COLOR_CARD, corner_radius=12, border_width=1, border_color=COLOR_PRIMARY)
            row.pack(fill="x", pady=8)

            # Icon and type
            icon = "🔼" if tx["type"] == "send" else "🔽"
            color = COLOR_DANGER if tx["type"] == "send" else COLOR_SUCCESS
            
            # Top row: Icon + type/token + amount
            top = ctk.CTkFrame(row, fg_color="transparent")
            top.pack(fill="x", padx=15, pady=(12, 8))

            ctk.CTkLabel(top, text=icon, font=("Arial", 20)).pack(side="left", padx=(0, 10))

            left = ctk.CTkFrame(top, fg_color="transparent")
            left.pack(side="left", fill="both", expand=True)
            
            action = f"{tx['type'].upper()} {tx['token']}"
            ctk.CTkLabel(left, text=action, font=("Inter", 13, "bold"), text_color=COLOR_TEXT).pack(anchor="w")
            ctk.CTkLabel(left, text=tx["status"].upper(), text_color=COLOR_SUCCESS if tx["status"] == "confirmed" else COLOR_TEXT_MUTED, 
                        font=("Inter", 10, "bold")).pack(anchor="w", pady=(2, 0))

            right = ctk.CTkFrame(top, fg_color="transparent")
            right.pack(side="right", padx=(10, 0), fill="y")

            sign = "-" if tx["type"] == "send" else "+"
            ctk.CTkLabel(right, text=f"{sign}{tx['amount']:.6g}", font=("Inter", 13, "bold"), text_color=color).pack(anchor="e")
            fee_text = f"Fee: {tx.get('fee', 0):.6g} {tx['token']}" if tx["type"] == "send" else ""
            if fee_text:
                ctk.CTkLabel(right, text=fee_text, text_color=COLOR_TEXT_MUTED, font=("Inter", 9)).pack(anchor="e", pady=(2, 0))

            # Bottom row: Recipient/Address info + Time
            bottom = ctk.CTkFrame(row, fg_color="transparent")
            bottom.pack(fill="x", padx=15, pady=(0, 12))

            # Recipient info
            recipient_info = f"Sent to: {tx.get('recipient', 'Unknown')[:20]}..." if tx["type"] == "send" else "Received"
            ctk.CTkLabel(bottom, text=recipient_info, text_color=COLOR_TEXT_MUTED, font=("Inter", 10)).pack(anchor="w")

            # Time
            time_str = tx["time"].strftime("%b %d, %H:%M") if hasattr(tx["time"], "strftime") else str(tx["time"])
            ctk.CTkLabel(bottom, text=time_str, text_color=COLOR_TEXT_MUTED, font=("Inter", 9)).pack(anchor="w", pady=(3, 0))

if __name__ == "__main__":
    app = App()
    app.mainloop()